#!/usr/bin/env python3
"""
Head-to-head drug comparison between old and new American Hospital Excel files.
Creates comparison tables and visualizations of changes per drug.
"""

import pandas as pd
import numpy as np
from pathlib import Path

try:
    import matplotlib.pyplot as plt
    import matplotlib
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

# Drug name normalization: map variants to canonical display names
DRUG_ALIASES = {
    "Other. (somatrogon (NGENLA) 60mg/1.2mL PFP 1's)": "Ngenla (Pfizer)",
    "Other. (somatrogon (NGENLA)": "Ngenla (Pfizer)",
    "Other, please specify": None,  # Exclude from comparison (generic placeholder)
    "Other": "Ngenla (Pfizer)",  # Details sheet
}

# Canonical list of GHD drugs (growth hormone deficiency medications) for calculators/forms
GHD_DRUGS = [
    "Genotropin (Pfizer)",
    "Norditropin (Novo Nordisk)",
    "Nutropin (Genentech)",
    "Omnitrope (Sandoz)",
    "Saizen (Merck Serono)",
    "Humatrope (Eli Lilly)",
    "Ngenla (Pfizer)",
    "Zomacton (Ferring)",
]


def normalize_drug_name(name: str) -> str | None:
    """Map drug name variants to canonical names for comparison."""
    s = str(name).strip()
    for alias, canonical in DRUG_ALIASES.items():
        if alias in s or s.startswith(alias[:20]):
            return canonical
    return s if s else None


def parse_american_hospital_excel(filepath: str) -> pd.DataFrame:
    """
    Parse American Hospital GHD Excel files (non-standard layout).
    Returns DataFrame with Drug as index and metric columns.
    """
    df_raw = pd.read_excel(filepath, header=None)
    
    # Find header row (row containing "Drug")
    header_row = None
    drug_col_idx = None
    for i in range(min(5, len(df_raw))):
        for j, val in enumerate(df_raw.iloc[i]):
            if "Drug" in str(val):
                header_row = i
                drug_col_idx = j
                break
        if header_row is not None:
            break
    
    if header_row is None:
        raise ValueError("Could not find header row with 'Drug'")
    
    # Extract headers (row after header_row may have time period - use header_row)
    headers = df_raw.iloc[header_row].tolist()
    
    # Get data rows (skip Total/summary rows)
    data_rows = []
    for i in range(header_row + 1, len(df_raw)):
        drug_name = df_raw.iloc[i, drug_col_idx]
        if pd.isna(drug_name):
            continue
        drug_str = str(drug_name).strip()
        # Skip Total/summary rows
        if "Total" in drug_str or "UNIQUE" in drug_str or drug_str == "":
            continue
        
        # Normalize drug names (e.g. "Other (NGENLA)" -> "Ngenla (Pfizer)")
        canonical = normalize_drug_name(drug_str)
        if canonical is None:
            continue  # Excluded alias (e.g. generic "Other, please specify")
        drug_str = canonical
        
        row = {"Drug": drug_str}
        for j, h in enumerate(headers):
            if j == drug_col_idx:
                continue
            val = df_raw.iloc[i, j]
            try:
                row[str(h)] = float(val) if pd.notna(val) else np.nan
            except (ValueError, TypeError):
                continue
        data_rows.append(row)
    
    if not data_rows:
        return pd.DataFrame()
    
    df = pd.DataFrame(data_rows).set_index("Drug")
    return df


def _normalize_sheet_name(s: str) -> str:
    """Normalize for comparison: strip and lower."""
    return str(s).strip().lower() if s else ""


def get_details_sheet_name(xl: pd.ExcelFile) -> str | None:
    """
    Detect the 'details' sheet: exact 'Questionnaire - details' (case-insensitive),
    or name containing 'details', or second sheet.
    Returns the actual sheet name as it appears in the file.
    """
    names = xl.sheet_names
    # 1. Exact or case-insensitive match for "Questionnaire - details"
    for n in names:
        if _normalize_sheet_name(n) == "questionnaire - details":
            return n
    if "Questionnaire - details" in names:
        return "Questionnaire - details"
    # 2. Any name containing "details"
    for n in names:
        if n and "details" in _normalize_sheet_name(n):
            return n
    # 3. Second sheet (index 1) if present
    if len(names) >= 2:
        return names[1]
    return None


def get_second_tab_sheet_names(old_path: str, new_path: str) -> tuple[str | None, str | None]:
    """
    Return (old_second_sheet_name, new_second_sheet_name) when both files have at least 2 sheets.
    Use this to compare the second tab in both files with the same instructions as the overview.
    """
    xl_old = pd.ExcelFile(old_path)
    xl_new = pd.ExcelFile(new_path)
    name_old = xl_old.sheet_names[1] if len(xl_old.sheet_names) >= 2 else None
    name_new = xl_new.sheet_names[1] if len(xl_new.sheet_names) >= 2 else None
    return (name_old, name_new)


def get_details_sheet_common(old_path: str, new_path: str) -> tuple[str | None, str | None]:
    """
    Return (sheet_name_for_old_file, sheet_name_for_new_file) for the details sheet.
    Each name is the exact string as it appears in that file, so parsing never fails on case mismatch.
    Returns (None, None) if no common details sheet is found.
    """
    xl_old = pd.ExcelFile(old_path)
    xl_new = pd.ExcelFile(new_path)
    name_old = get_details_sheet_name(xl_old)
    name_new = get_details_sheet_name(xl_new)
    if name_old and name_new and name_old == name_new:
        return (name_old, name_new)
    norm_old = _normalize_sheet_name(name_old) if name_old else ""
    norm_new = _normalize_sheet_name(name_new) if name_new else ""
    if norm_old and norm_new and norm_old == norm_new:
        return (name_old, name_new)
    for n in xl_new.sheet_names:
        if _normalize_sheet_name(n) == norm_old:
            return (name_old, n)
    for n in xl_old.sheet_names:
        if _normalize_sheet_name(n) == norm_new:
            return (n, name_new)
    if name_old and name_old in xl_new.sheet_names:
        return (name_old, name_old)
    if name_new and name_new in xl_old.sheet_names:
        return (name_new, name_new)
    return (None, None)


def _is_numeric_cell(val) -> bool:
    """True if value can be interpreted as a number (int or float)."""
    if pd.isna(val):
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val).strip())
        return True
    except (ValueError, TypeError):
        return False


def _looks_like_indication_section(type_str: str) -> bool:
    """True if type_str looks like an indication section header (e.g. pGHD, SGA, Turner Syndrome), not a drug."""
    s = type_str.strip().lower()
    if not s:
        return False
    # Subheader / skip row
    if "total patients" in s or "unique" in s:
        return True
    # Real section headers
    if "growth hormone" in s and ("def" in s or "ghd" in s):
        return True
    if "gestational" in s or "sga" in s:
        return True
    if "syndrome" in s and "turner" in s:
        return True
    return False




def parse_details_sheet(filepath: str, sheet_name: str = "Questionnaire - details") -> pd.DataFrame:
    """
    Parse the 'Questionnaire - details' sheet (indication × type/drug with metrics).
    Returns DataFrame with MultiIndex (Indication, Type) and metric columns.
    Tolerates 'Type' or 'Drug' header, flexible header row position, and multiple layout variants.
    """
    df_raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    if df_raw.empty or len(df_raw) < 2:
        return pd.DataFrame()
    nrows, ncols = len(df_raw), len(df_raw.columns)
    # Find header row: row containing "Type" or "Drug" (case-insensitive) in first columns
    header_row = None
    type_col_idx = 0
    header_keywords = ("type", "drug", "product", "treatment")
    for i in range(min(25, nrows)):
        for j in range(min(10, ncols)):
            cell = str(df_raw.iloc[i, j]).strip().lower()
            if any(kw in cell for kw in header_keywords) and len(cell) < 50:
                header_row = i
                type_col_idx = j
                break
        if header_row is not None:
            break
    if header_row is None:
        return pd.DataFrame()
    headers = [str(h).strip() if pd.notna(h) else "" for h in df_raw.iloc[header_row].tolist()]
    current_indication = None
    rows = []
    skip_type_values = {"scope", "time period", "type", "drug", "questionnaire", ""}
    for i in range(header_row + 1, nrows):
        type_val = df_raw.iloc[i, type_col_idx] if type_col_idx < ncols else np.nan
        if pd.isna(type_val):
            continue
        type_str = str(type_val).strip()
        if not type_str:
            continue
        if type_str.lower() in skip_type_values or "time period" in type_str.lower() or "scope" in type_str.lower():
            continue
        # Data row: first metric cell is numeric, OR any metric cell in row is numeric
        first_metric = df_raw.iloc[i, type_col_idx + 1] if type_col_idx + 1 < ncols else np.nan
        has_numeric = _is_numeric_cell(first_metric)
        if not has_numeric and type_col_idx + 1 < ncols:
            for j in range(type_col_idx + 1, min(type_col_idx + 6, ncols)):
                if _is_numeric_cell(df_raw.iloc[i, j]):
                    has_numeric = True
                    break
        is_data_row = has_numeric
        # Accept drug-name rows even with all NaN metrics (sheet may be unfilled; we emit 0.0s)
        if not is_data_row and current_indication and type_str != "Total patients (UNIQUE)" and type_str != "Other, please specify":
            if not _looks_like_indication_section(type_str):
                is_data_row = True
        if is_data_row:
            if current_indication is None or (isinstance(current_indication, float) and pd.isna(current_indication)):
                continue
            canonical_type = normalize_drug_name(type_str)
            if canonical_type is None and type_str == "Other, please specify":
                continue
            type_str = canonical_type if canonical_type else type_str
            row = {"Indication": current_indication, "Type": type_str}
            for j in range(type_col_idx + 1, len(headers)):
                if j >= ncols:
                    break
                h = headers[j]
                if not h:
                    continue
                val = df_raw.iloc[i, j]
                try:
                    row[h] = float(val) if pd.notna(val) else 0.0
                except (ValueError, TypeError):
                    row[h] = 0.0
            rows.append(row)
        else:
            # Section header: only set indication for real sections (pGHD, SGA, Turner), not Total/UNIQUE
            if type_str and _looks_like_indication_section(type_str) and "total patients" not in type_str.lower():
                current_indication = type_str

    # Fallback: if no header found or no rows, try first row as header and first column as Type
    if not rows and nrows >= 2 and ncols >= 2:
        header_row_fb = 0
        type_col_idx_fb = 0
        headers_fb = [str(h).strip() if pd.notna(h) else "" for h in df_raw.iloc[header_row_fb].tolist()]
        current_indication_fb = "Details"
        for i in range(1, nrows):
            type_val = df_raw.iloc[i, type_col_idx_fb] if type_col_idx_fb < ncols else np.nan
            if pd.isna(type_val):
                continue
            type_str = str(type_val).strip()
            if not type_str or type_str.lower() in skip_type_values:
                continue
            has_any_num = False
            for j in range(type_col_idx_fb + 1, ncols):
                if _is_numeric_cell(df_raw.iloc[i, j]):
                    has_any_num = True
                    break
            if not has_any_num:
                if not _is_numeric_cell(type_val) and "patients" not in type_str.lower() and not type_str.startswith("Total"):
                    current_indication_fb = type_str
                continue
            canonical_type = normalize_drug_name(type_str)
            if canonical_type is None and type_str == "Other, please specify":
                continue
            type_str = canonical_type if canonical_type else type_str
            row = {"Indication": current_indication_fb, "Type": type_str}
            for j in range(type_col_idx_fb + 1, len(headers_fb)):
                if j >= ncols:
                    break
                h = headers_fb[j]
                if not h:
                    continue
                val = df_raw.iloc[i, j]
                try:
                    row[h] = float(val) if pd.notna(val) else 0.0
                except (ValueError, TypeError):
                    row[h] = 0.0
            rows.append(row)

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df = df.set_index(["Indication", "Type"]).sort_index()
    return df


def get_details_sheet_row_col_map(
    filepath: str,
    sheet_name: str | None = None,
) -> tuple[dict, dict, int, int]:
    """
    Scan the details sheet and return mappings from (Indication, Type) to Excel row index
    and from metric header name to Excel column index (0-based).
    Uses the same parsing logic as parse_details_sheet so row/col indices match the sheet layout.
    Returns (row_map, col_map, header_row, type_col_idx).
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name is not None:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[1]
    # Build a raw grid (openpyxl is 1-based)
    max_row = ws.max_row
    max_col = ws.max_column or 10
    grid = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row.append(cell.value)
        grid.append(row)
    wb.close()

    df_raw = pd.DataFrame(grid)
    if df_raw.empty or len(df_raw) < 2:
        return {}, {}, 0, 0

    nrows, ncols = len(df_raw), len(df_raw.columns)
    header_row = None
    type_col_idx = 0
    header_keywords = ("type", "drug", "product", "treatment")
    for i in range(min(25, nrows)):
        for j in range(min(10, ncols)):
            cell = str(df_raw.iloc[i, j]).strip().lower()
            if any(kw in cell for kw in header_keywords) and len(cell) < 50:
                header_row = i
                type_col_idx = j
                break
        if header_row is not None:
            break
    if header_row is None:
        return {}, {}, 0, 0

    headers = [str(h).strip() if pd.notna(h) else "" for h in df_raw.iloc[header_row].tolist()]
    row_map = {}
    col_map = {}
    for j in range(type_col_idx + 1, len(headers)):
        h = headers[j] if j < len(headers) else ""
        if h:
            col_map[h] = j

    current_indication = None
    skip_type_values = {"scope", "time period", "type", "drug", "questionnaire", ""}
    for i in range(header_row + 1, nrows):
        type_val = df_raw.iloc[i, type_col_idx] if type_col_idx < ncols else np.nan
        if pd.isna(type_val):
            continue
        type_str = str(type_val).strip()
        if not type_str or type_str.lower() in skip_type_values or "time period" in type_str.lower() or "scope" in type_str.lower():
            continue
        first_metric = df_raw.iloc[i, type_col_idx + 1] if type_col_idx + 1 < ncols else np.nan
        has_numeric = _is_numeric_cell(first_metric)
        if not has_numeric and type_col_idx + 1 < ncols:
            for j in range(type_col_idx + 1, min(type_col_idx + 6, ncols)):
                if _is_numeric_cell(df_raw.iloc[i, j]):
                    has_numeric = True
                    break
        is_data_row = has_numeric
        if not is_data_row and current_indication and type_str != "Total patients (UNIQUE)" and type_str != "Other, please specify":
            if not _looks_like_indication_section(type_str):
                is_data_row = True
        if is_data_row:
            if current_indication is None or (isinstance(current_indication, float) and pd.isna(current_indication)):
                continue
            canonical_type = normalize_drug_name(type_str)
            if canonical_type is None and type_str == "Other, please specify":
                continue
            type_key = canonical_type if canonical_type else type_str
            row_map[(current_indication, type_key)] = i
        else:
            if type_str and _looks_like_indication_section(type_str) and "total patients" not in type_str.lower():
                current_indication = type_str

    return row_map, col_map, header_row, type_col_idx


# Column name mapping: Old details (12-month) -> New sheet (6-month) for copying Old to New
DETAILS_12MO_TO_6MO_COLUMNS = {
    "Total patients treated in the last 12 months": "Total patients treated in the last 6 months",
    "Newly diagnosed patients received drug treatments in the last 12 months": "Newly diagnosed patients received drug treatments in the last 6 months",
    "Follow up patients and received drug treatments in the last 12 months": "Follow up patients and received drug treatments in the last 6 months",
    "All newly diagnosed in the last 12 months": "All newly diagnosed in the last 6 months",
    "Newly diagnosed patients but didn't receive drug treatments in the last 12 months": "Newly diagnosed patients but didn't receive drug treatments in the last 6 months",
    "All follow up patients (12 mo)": "All follow up patients",
    "Follow up patients and didn't receive drug treatments in the last 12 months": "Follow up patients and didn't receive drug treatments in the last 6 months",
    "Active patients on treatments (Received treatments in April - June 2025)": "Active patients (Received treatments in October - December 2025)",
}


def details_old_to_new_columns(old_details_df: pd.DataFrame) -> pd.DataFrame:
    """
    Return a copy of old_details (12-month columns) with column names mapped to 6-month
    so it can be written to the New American second sheet via fill_details_sheet_to_excel.
    """
    if old_details_df.empty:
        return old_details_df.copy()
    rename_map = {}
    for c in old_details_df.columns:
        if c in DETAILS_12MO_TO_6MO_COLUMNS:
            rename_map[c] = DETAILS_12MO_TO_6MO_COLUMNS[c]
        elif isinstance(c, str) and "12 months" in c:
            rename_map[c] = c.replace("12 months", "6 months")
    out = old_details_df.rename(columns=rename_map)
    return out


def fill_details_sheet_to_excel(
    filepath: str,
    details_df: pd.DataFrame,
    sheet_name: str | None = None,
) -> int:
    """
    Fill the details DataFrame into the second sheet of the Excel file.
    Matches (Indication, Type) to rows and metric column names to columns;
    writes each value into the correct cell for that drug/section.
    Returns the number of cells written.
    Saves via a temp file then replaces the original so changes persist on disk.
    """
    import openpyxl
    import tempfile
    import os

    if details_df.empty or not details_df.index.names or list(details_df.index.names) != ["Indication", "Type"]:
        raise ValueError("details_df must have MultiIndex (Indication, Type)")
    row_map, col_map, _header_row, _type_col_idx = get_details_sheet_row_col_map(filepath, sheet_name)
    if not row_map or not col_map:
        raise ValueError("Could not find details layout in the sheet (header row with 'Type', data rows)")

    wb = openpyxl.load_workbook(filepath)
    if sheet_name is not None:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[1]

    # Normalize col_map keys for matching (strip)
    col_map_normalized = {str(k).strip(): v for k, v in col_map.items()}
    df_cols_stripped = [str(c).strip() for c in details_df.columns]
    cols_matched = [c for c in df_cols_stripped if c in col_map_normalized]

    def _row_key_for(key: tuple) -> int | None:
        """Return 0-based row index in sheet for (Indication, Type), or None if no match."""
        ind, typ = key
        if (ind, typ) in row_map:
            return row_map[(ind, typ)]
        # Fuzzy: match by Type and indication substring (e.g. "pGHD" in "Pediatric ... (pGHD)")
        ind_s, typ_s = str(ind).strip(), str(typ).strip()
        for (sheet_ind, sheet_typ), ri in row_map.items():
            if str(sheet_typ).strip() != typ_s:
                continue
            si = str(sheet_ind).strip()
            if si in ind_s or ind_s in si:
                return ri
        return None

    # Fuzzy column match: normalize spaces so "6  months" or trailing space still matches
    def _norm(s: str) -> str:
        return " ".join(str(s).strip().split())

    col_map_fuzzy = {_norm(k): (k, v) for k, v in col_map.items()}

    written = 0
    for key in details_df.index:
        row_i = _row_key_for(key)
        if row_i is None:
            continue
        row_idx_1based = row_i + 1
        for col_name in details_df.columns:
            col_stripped = str(col_name).strip()
            col_idx_0 = None
            if col_stripped in col_map_normalized:
                col_idx_0 = col_map_normalized[col_stripped]
            elif _norm(col_name) in col_map_fuzzy:
                _, col_idx_0 = col_map_fuzzy[_norm(col_name)]
            if col_idx_0 is None:
                continue
            col_idx_1based = col_idx_0 + 1
            try:
                val = details_df.loc[key, col_name]
                if pd.isna(val):
                    val = 0
                else:
                    val = float(val)
            except (TypeError, ValueError):
                val = 0
            ws.cell(row=row_idx_1based, column=col_idx_1based, value=val)
            written += 1

    if written == 0:
        keys_in_sheet = set(row_map.keys())
        keys_in_df = set(details_df.index)
        sample_sheet = list(keys_in_sheet)[:3] if keys_in_sheet else []
        sample_df = list(keys_in_df)[:3] if keys_in_df else []
        sheet_cols = list(col_map_normalized.keys())[:5]
        raise ValueError(
            "No cells written: row or column mismatch. "
            f"Data has {len(keys_in_df)} (Indication,Type) rows, sheet map has {len(keys_in_sheet)}. "
            f"Data columns matched: {len(cols_matched)} of {len(details_df.columns)} (sheet has {len(col_map_normalized)} metric cols). "
            f"Sample sheet rows: {sample_sheet}. Sample data rows: {sample_df}. Sample sheet cols: {sheet_cols}."
        )

    # Save to temp file in same directory, then replace original (ensures persistence)
    abs_path = os.path.abspath(filepath)
    dirpath = os.path.dirname(abs_path) or "."
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=dirpath)
    try:
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, abs_path)
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        raise
    return written


def compare_details(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """
    Head-to-head comparison for details sheet: align by (Indication, Type), compare metrics.
    """
    if old_df.empty and new_df.empty:
        return pd.DataFrame()
    # Align columns (old has 12mo, new has 6mo labels)
    metric_keys = [
        ("Total patients treated in the last 12 months", "Total patients treated in the last 6 months"),
        ("Newly diagnosed patients received drug treatments in the last 12 months", "Newly diagnosed patients received drug treatments in the last 6 months"),
        ("Follow up patients and received drug treatments in the last 12 months", "Follow up patients and received drug treatments in the last 6 months"),
    ]
    col_pairs = []
    for ok, nk in metric_keys:
        if ok in old_df.columns and nk in new_df.columns:
            col_pairs.append((ok, nk))
    if not col_pairs:
        # Fallback: match by position
        numeric_old = [c for c in old_df.columns if pd.api.types.is_numeric_dtype(old_df[c])][:4]
        numeric_new = [c for c in new_df.columns if pd.api.types.is_numeric_dtype(new_df[c])][:4]
        for i in range(min(len(numeric_old), len(numeric_new))):
            col_pairs.append((numeric_old[i], numeric_new[i]))
    all_keys = sorted(set(old_df.index) | set(new_df.index))
    out_rows = []
    for key in all_keys:
        ind, typ = key
        row_data = {"Indication": ind, "Type": typ}
        for old_col, new_col in col_pairs:
            metric = str(old_col)[:35].replace("12 months", "12mo").replace("6 months", "6mo")
            try:
                old_val = float(old_df.at[key, old_col]) if key in old_df.index and old_col in old_df.columns else 0.0
                new_val = float(new_df.at[key, new_col]) if key in new_df.index and new_col in new_df.columns else 0.0
            except (KeyError, TypeError, ValueError):
                old_val, new_val = 0.0, 0.0
            if pd.isna(old_val):
                old_val = 0.0
            if pd.isna(new_val):
                new_val = 0.0
            row_data[f"{metric}_old"] = old_val
            row_data[f"{metric}_new"] = new_val
            row_data[f"{metric}_change"] = new_val - old_val
        out_rows.append(row_data)
    df = pd.DataFrame(out_rows)
    df = df.set_index(["Indication", "Type"])
    return df


def compare_drugs(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """
    Head-to-head comparison: align drugs and compare matching metrics.
    """
    # Find matching columns (same or similar metric names)
    old_cols = list(old_df.columns)
    new_cols = list(new_df.columns)
    
    # Map by position/keyword - common metrics between files
    metric_keys = [
        ("Total", "Total"),
        ("Newly diagnosed patients received", "Newly diagnosed"),
        ("Follow up patients and received", "Follow up received"),
        ("Active patients", "Active patients"),
    ]
    
    col_pairs = []
    for ok, nk in metric_keys:
        for oc in old_cols:
            if ok in str(oc):
                for nc in new_cols:
                    if nk in str(nc) or (ok in str(nc)):
                        col_pairs.append((oc, nc))
                        break
                break
    
    # If no keyword match, use positional: first N numeric columns
    if not col_pairs:
        numeric_old = [c for c in old_df.columns if pd.api.types.is_numeric_dtype(old_df[c])]
        numeric_new = [c for c in new_df.columns if pd.api.types.is_numeric_dtype(new_df[c])]
        if not numeric_old:
            numeric_old = list(old_df.columns)[:5]
        if not numeric_new:
            numeric_new = list(new_df.columns)[:5]
        for i in range(min(4, len(numeric_old), len(numeric_new))):
            col_pairs.append((numeric_old[i], numeric_new[i]))
    
    # All drugs (union)
    all_drugs = sorted(set(old_df.index) | set(new_df.index))
    
    # Build comparison
    rows = []
    for drug in all_drugs:
        row_data = {"Drug": drug}
        for old_col, new_col in col_pairs:
            if old_col not in old_df.columns or new_col not in new_df.columns:
                continue
            metric = str(old_col)[:40].replace("Total patients treated in the last 12 months", "Total").replace("Total patients treated in the last 6 months", "Total")
            old_val = old_df.loc[drug, old_col] if drug in old_df.index else np.nan
            new_val = new_df.loc[drug, new_col] if drug in new_df.index else np.nan
            old_val = float(old_val) if pd.notna(old_val) else 0.0
            new_val = float(new_val) if pd.notna(new_val) else 0.0
            
            row_data[f"{metric}_old"] = old_val
            row_data[f"{metric}_new"] = new_val
            row_data[f"{metric}_change"] = new_val - old_val
        
        rows.append(row_data)
    
    return pd.DataFrame(rows).set_index("Drug")


# McKinsey Elite style palette (exported for app.py)
MCKINSEY_STYLE = {
    "navy": "#1a365d",
    "charcoal": "#2d3748",
    "slate": "#4a5568",
    "teal": "#2c7a7b",
    "accent_blue": "#3182ce",
    "accent_teal": "#38a169",
    "decrease": "#c53030",
    "increase": "#276749",
    "bg": "#ffffff",
    "grid": "#e2e8f0",
    "text": "#2d3748",
}


def _mckinsey_style(ax, title: str = ""):
    """Apply McKinsey elite styling to matplotlib axes."""
    M = MCKINSEY_STYLE
    ax.set_facecolor(M["bg"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color(M["grid"])
    ax.spines["bottom"].set_color(M["grid"])
    ax.tick_params(colors=M["text"], labelsize=10)
    ax.yaxis.grid(True, color=M["grid"], linestyle="-", linewidth=0.8)
    ax.set_axisbelow(True)
    if title:
        ax.set_title(title, fontsize=12, fontweight=600, color=M["charcoal"], pad=12)


def create_visualizations(comp_df: pd.DataFrame, output_dir: str = "."):
    """Create McKinsey-style bar charts: old vs new per drug, and change per drug."""
    if not HAS_MATPLOTLIB:
        print("Skipping visualizations: matplotlib not installed. Run: pip install matplotlib")
        return
    output_dir = Path(output_dir)
    matplotlib.rcParams["font.family"] = "sans-serif"
    matplotlib.rcParams["font.sans-serif"] = ["Helvetica Neue", "Arial", "DejaVu Sans"]
    
    # Get metric columns
    old_cols = [c for c in comp_df.columns if c.endswith("_old")]
    new_cols = [c for c in comp_df.columns if c.endswith("_new")]
    
    # Match pairs
    metrics = []
    for oc in old_cols:
        base = oc.replace("_old", "")
        nc = f"{base}_new"
        if nc in comp_df.columns:
            metrics.append((base, oc, nc))
    
    if not metrics:
        return
    
    drugs = [d for d in comp_df.index if comp_df.loc[d, old_cols + new_cols].notna().any()]
    if not drugs:
        drugs = comp_df.index.tolist()
    
    def short_name(d):
        s = str(d).replace("(Pfizer)", "").replace("(Novo Nordisk)", "").replace("(Merck)", "").replace("(Sandoz)", "").strip()
        return s[:20] if len(s) > 20 else s
    
    drug_labels = [short_name(d) for d in drugs]
    
    # Chart 1: Old vs New side-by-side (McKinsey style)
    total_old = next((c for c in old_cols if "Total" in c), old_cols[0] if old_cols else None)
    total_new = next((c for c in new_cols if "Total" in c), new_cols[0] if new_cols else None)
    
    if total_old and total_new:
        fig, ax = plt.subplots(figsize=(12, 6))
        fig.patch.set_facecolor(MCKINSEY_STYLE["bg"])
        x = np.arange(len(drugs))
        w = 0.36
        
        old_vals = [comp_df.loc[d, total_old] if d in comp_df.index else 0 for d in drugs]
        new_vals = [comp_df.loc[d, total_new] if d in comp_df.index else 0 for d in drugs]
        old_vals = [v if pd.notna(v) else 0 for v in old_vals]
        new_vals = [v if pd.notna(v) else 0 for v in new_vals]
        
        ax.bar(x - w/2, old_vals, w, label="Prior period (12 mo)", color=MCKINSEY_STYLE["navy"], alpha=0.92)
        ax.bar(x + w/2, new_vals, w, label="Current period (6 mo)", color=MCKINSEY_STYLE["teal"], alpha=0.92)
        ax.set_ylabel("Patients", fontsize=11, color=MCKINSEY_STYLE["text"])
        _mckinsey_style(ax, "Total patients by drug — prior vs current period")
        ax.set_xticks(x)
        ax.set_xticklabels(drug_labels, rotation=45, ha="right", fontsize=10)
        ax.legend(frameon=False, fontsize=10, loc="upper right")
        plt.tight_layout()
        plt.savefig(output_dir / "drug_total_comparison.png", dpi=150, bbox_inches="tight", facecolor=MCKINSEY_STYLE["bg"])
        plt.close()
    
    # Chart 2: Change (delta) per drug
    change_cols = [c for c in comp_df.columns if c.endswith("_change")]
    if change_cols:
        ch_col = next((c for c in change_cols if "Total" in c), change_cols[0])
        changes = [comp_df.loc[d, ch_col] if d in comp_df.index else np.nan for d in drugs]
        changes = [c if pd.notna(c) else 0 for c in changes]
        
        fig2, ax2 = plt.subplots(figsize=(10, 6))
        fig2.patch.set_facecolor(MCKINSEY_STYLE["bg"])
        colors = [MCKINSEY_STYLE["increase"] if c >= 0 else MCKINSEY_STYLE["decrease"] for c in changes]
        ax2.barh(drug_labels, changes, color=colors, alpha=0.9, height=0.6)
        ax2.axvline(x=0, color=MCKINSEY_STYLE["charcoal"], linewidth=1.2)
        ax2.set_xlabel("Δ Patients (current − prior)", fontsize=11, color=MCKINSEY_STYLE["text"])
        _mckinsey_style(ax2, "Change in total patients by drug")
        ax2.xaxis.grid(True, color=MCKINSEY_STYLE["grid"], linestyle="-", linewidth=0.8)
        plt.tight_layout()
        plt.savefig(output_dir / "drug_change_by_drug.png", dpi=150, bbox_inches="tight", facecolor=MCKINSEY_STYLE["bg"])
        plt.close()
    
    # Chart 3: Multi-metric comparison
    if len(metrics) >= 2:
        fig3, axes = plt.subplots(2, 2, figsize=(16, 11))
        fig3.patch.set_facecolor(MCKINSEY_STYLE["bg"])
        axes = axes.flatten()
        for idx, (base, oc, nc) in enumerate(metrics[:4]):
            ax = axes[idx]
            old_vals = [comp_df.loc[d, oc] if d in comp_df.index else 0 for d in drugs]
            new_vals = [comp_df.loc[d, nc] if d in comp_df.index else 0 for d in drugs]
            old_vals = [v if pd.notna(v) else 0 for v in old_vals]
            new_vals = [v if pd.notna(v) else 0 for v in new_vals]
            
            x = np.arange(len(drugs))
            w = 0.36
            ax.bar(x - w/2, old_vals, w, label="Prior", color=MCKINSEY_STYLE["navy"], alpha=0.92)
            ax.bar(x + w/2, new_vals, w, label="Current", color=MCKINSEY_STYLE["teal"], alpha=0.92)
            _mckinsey_style(ax, str(base).replace("Total patients treated in the last 12 months", "Total patients")[:45])
            ax.set_xticks(x)
            ax.set_xticklabels(drug_labels, rotation=45, ha="right", fontsize=9)
            ax.legend(frameon=False, fontsize=9, loc="upper left")
        # Increase spacing between subplots
        plt.subplots_adjust(hspace=0.35, wspace=0.25)
        plt.savefig(output_dir / "drug_metrics_comparison.png", dpi=150, bbox_inches="tight", facecolor=MCKINSEY_STYLE["bg"])
        plt.close()


def main():
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python drug_comparison.py <old_file.xlsx> <new_file.xlsx>")
        sys.exit(1)
    
    old_path = sys.argv[1]
    new_path = sys.argv[2]
    
    if not Path(old_path).exists():
        print(f"Error: Old file not found: {old_path}")
        sys.exit(1)
    if not Path(new_path).exists():
        print(f"Error: New file not found: {new_path}")
        sys.exit(1)
    
    print("Parsing Excel files (both sheets)...")
    # Sheet 1: Questionnaire - overview (drug-level)
    old_overview = parse_american_hospital_excel(old_path)
    new_overview = parse_american_hospital_excel(new_path)
    comp_overview = compare_drugs(old_overview, new_overview)
    
    # Sheet 2: details (detect by name "details" or second tab)
    old_details = pd.DataFrame()
    new_details = pd.DataFrame()
    comp_details = pd.DataFrame()
    details_name_old, details_name_new = get_details_sheet_common(old_path, new_path)
    if details_name_old and details_name_new:
        old_details = parse_details_sheet(old_path, details_name_old)
        new_details = parse_details_sheet(new_path, details_name_new)
        comp_details = compare_details(old_details, new_details)
        print(f"\nDetails sheet: \"{details_name_old}\" / \"{details_name_new}\" — Old rows: {len(old_details)}, New rows: {len(new_details)}, Comparison rows: {len(comp_details)}")
    
    print(f"\nOverview - Old drugs: {list(old_overview.index)}")
    print(f"Overview - New drugs: {list(new_overview.index)}")
    
    # Save Excel with both sheet comparisons (NaN → 0)
    output_excel = "drug_comparison_head_to_head.xlsx"
    with pd.ExcelWriter(output_excel, engine="xlsxwriter") as writer:
        comp_overview.fillna(0).to_excel(writer, sheet_name="Overview Comparison")
        comp_details.fillna(0).to_excel(writer, sheet_name="Details Comparison")
        old_overview.fillna(0).to_excel(writer, sheet_name="Overview Prior")
        new_overview.fillna(0).to_excel(writer, sheet_name="Overview Current")
        if not old_details.empty:
            old_details.fillna(0).to_excel(writer, sheet_name="Details Prior")
        if not new_details.empty:
            new_details.fillna(0).to_excel(writer, sheet_name="Details Current")
        
        header_fmt = writer.book.add_format({
            "bold": True, "font_size": 11, "font_color": "#1a365d",
            "bottom": 2, "bottom_color": "#2d3748",
        })
        for sheet_name, ws in writer.sheets.items():
            ws.set_column("A:A", 32)
            ws.set_column("B:Z", 14)
    
    print(f"\nComparisons saved to: {output_excel}")
    print("  Sheets: Overview Comparison, Details Comparison, Overview Prior/Current, Details Prior/Current")
    
    # Visualizations (overview only)
    print("Creating visualizations (overview)...")
    create_visualizations(comp_overview)
    if HAS_MATPLOTLIB:
        print("Charts saved: drug_total_comparison.png, drug_change_by_drug.png, drug_metrics_comparison.png")
    
    print("\n" + "=" * 70)
    print("OVERVIEW COMPARISON (Head-to-Head)")
    print("=" * 70)
    print(comp_overview.to_string())
    if not comp_details.empty:
        print("\n" + "=" * 70)
        print("DETAILS COMPARISON (Indication × Type)")
        print("=" * 70)
        print(comp_details.to_string())
    print("=" * 70)


if __name__ == "__main__":
    main()
